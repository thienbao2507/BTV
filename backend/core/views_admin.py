# core/views_admin.py

import os
import csv
import re
import unicodedata
from io import TextIOWrapper

from django.conf import settings
from django.contrib import messages
from django.core.files.storage import default_storage
from django.core.files.uploadedfile import InMemoryUploadedFile, TemporaryUploadedFile
from django.db import transaction
from django.db.models import Prefetch
from django.shortcuts import render, redirect

from openpyxl import load_workbook

from core.decorators import judge_required
from .models import (
    BaiThi,
    BaiThiTemplateSection,
    BaiThiTemplateItem,
    VongThi,
    ThiSinh,
    GiamKhao,
    CuocThi,
    ThiSinhCuocThi,
)

# ============================================================
# CẤU HÌNH CỘT HỖ TRỢ IMPORT
# ============================================================

REQUIRED_COLUMNS = {
    "thisinh": ["maNV", "hoTen", "chiNhanh", "vung", "donVi", "email", "nhom", "image_url"],
    "giamkhao": ["maNV", "hoTen", "email"],
    "voting":  ["maNV", "hoTen", "chiNhanh", "vung", "donVi", "email", "nhom", "image_url"],
}

def _normalize(s: str) -> str:
    s = unicodedata.normalize("NFKD", s)
    s = s.encode("ascii", "ignore").decode("ascii")
    s = s.lower()
    s = re.sub(r"[^a-z0-9]+", "", s)
    return s

HEADER_ALIASES = {
    # maNV
    "manv": "maNV", "manhanvien": "maNV", "manvnv": "maNV", "ma": "maNV", "ma_nv": "maNV",
    # hoTen
    "hoten": "hoTen", "ten": "hoTen", "hovaten": "hoTen", "ho_ten": "hoTen",
    # chiNhanh
    "chinhanh": "chiNhanh", "chi_nhanh": "chiNhanh", "cn": "chiNhanh",
    # vung
    "vung": "vung", "mien": "vung",
    # donVi
    "donvi": "donVi", "don_vi": "donVi", "dv": "donVi", "don": "donVi", "donvichuyendoi": "donVi",
    # email
    "email": "email", "mail": "email", "e-mail": "email",
    # nhom
    "nhom": "nhom", "group": "nhom", "nhomthi": "nhom",
    # image_url
    "imageurl": "image_url", "image_url": "image_url", "hinhanh": "image_url",
    "hinh_anh": "image_url", "hinhanh": "image_url", "anh": "image_url", "img": "image_url",
}

def _map_header_list(header, expected_cols):
    """
    Trả về: (canon_order, src_idx, missing)
    - canon_order: danh sách tên cột đã map về canonical theo thứ tự header gốc
    - src_idx: dict {canonical_name: index_goc}
    - missing: các cột expected chưa có
    """
    canon_order = []
    for h in header:
        key = _normalize(h or "")
        canon = HEADER_ALIASES.get(key)
        canon_order.append(canon or (h or "").strip())

    src_idx = {}
    for i, canon in enumerate(canon_order):
        if canon not in src_idx:
            src_idx[canon] = i

    missing = [c for c in expected_cols if c not in src_idx]
    return canon_order, src_idx, missing

def _read_xlsx(file, expected_cols):
    wb = load_workbook(file, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(c).strip() if c is not None else "" for c in rows[0]]

    _, src_idx, missing = _map_header_list(header, expected_cols)
    if missing:
        raise ValueError(f"Thiếu cột: {', '.join(missing)}")

    data = []
    for r in rows[1:]:
        if r is None:
            continue
        row = {}
        for c in expected_cols:
            idx = src_idx[c]
            val = r[idx] if idx < len(r) else None
            row[c] = "" if val is None else str(val).strip()
        data.append(row)
    return data

def _read_csv(file, expected_cols):
    # UploadedFile -> decode UTF-8
    text_stream = TextIOWrapper(file, encoding="utf-8")
    reader = csv.DictReader(text_stream)
    header = reader.fieldnames or []

    _, src_idx, missing = _map_header_list(header, expected_cols)
    if missing:
        raise ValueError(f"Thiếu cột: {', '.join(missing)}")

    # map canonical -> tên cột gốc
    canon_to_source = {}
    for h in header:
        key = _normalize(h or "")
        canon = HEADER_ALIASES.get(key) or (h or "").strip()
        if canon not in canon_to_source:
            canon_to_source[canon] = h

    data = []
    for row in reader:
        out = {}
        for c in expected_cols:
            src = canon_to_source.get(c, c)
            out[c] = (row.get(src, "") or "").strip()
        data.append(out)
    return data

def _find_duplicate_ma_email(rows, key_ma="maNV", key_email="email"):
    seen_ma, seen_email = set(), set()
    dup_ma, dup_email = set(), set()
    for r in rows:
        ma = (r.get(key_ma) or "").strip()
        if ma:
            if ma in seen_ma:
                dup_ma.add(ma)
            else:
                seen_ma.add(ma)
        email = (r.get(key_email) or "").strip().lower()
        if email:
            if email in seen_email:
                dup_email.add(email)
            else:
                seen_email.add(email)
    return dup_ma, dup_email

# ============================================================
# IMPORT THÍ SINH/GIÁM KHẢO/VOTING
# ============================================================

@judge_required
def import_view(request):
    """
    - Chọn cuộc thi (maCT)
    - target = thisinh | giamkhao | voting
    - File CSV/XLSX
    * Với 'voting': sau khi tạo/cập nhật ThiSinh + ThiSinhCuocThi, sẽ tự thêm vào ThiSinhVoting.
    """
    preselected_ma = None
    q = request.GET.get("ct") or request.POST.get("maCT")
    if q:
        if CuocThi.objects.filter(ma=q).exists():
            preselected_ma = q
        else:
            try:
                obj = CuocThi.objects.only("ma").get(pk=int(q))
                preselected_ma = obj.ma
            except Exception:
                pass

    if request.method == "POST":
        target = (request.POST.get("target") or "").strip()  # thisinh | giamkhao | voting
        selected_ma_ct = (request.POST.get("maCT") or "").strip()
        f = request.FILES.get("file")

        cuocthi_obj = CuocThi.objects.filter(ma=selected_ma_ct).first() if selected_ma_ct else None

        if target not in REQUIRED_COLUMNS:
            messages.error(request, "Vui lòng chọn loại dữ liệu hợp lệ.")
            return redirect(request.path)
        if not f:
            messages.error(request, "Vui lòng chọn tệp CSV/XLSX.")
            return redirect(request.path)

        expected = REQUIRED_COLUMNS[target]
        try:
            if isinstance(f, (InMemoryUploadedFile, TemporaryUploadedFile)) and f.name.lower().endswith(".xlsx"):
                rows = _read_xlsx(f, expected)
            else:
                rows = _read_csv(f, expected)
        except Exception as e:
            messages.error(request, f"Lỗi đọc tệp: {e}")
            return redirect(request.path)

        # Chặn trùng trong chính file import
        dup_ma, dup_email = _find_duplicate_ma_email(rows)
        if dup_ma or dup_email:
            loai = "thí sinh" if target in ("thisinh", "voting") else "giám khảo"
            parts = []
            if dup_ma:
                parts.append("Mã nhân viên trùng: " + ", ".join(sorted(dup_ma)))
            if dup_email:
                parts.append("Email trùng: " + ", ".join(sorted(dup_email)))
            prefix = f"Không thể import {loai}"
            if cuocthi_obj:
                prefix += f" vào cuộc thi {cuocthi_obj.ma}"
            messages.error(request, prefix + ". " + " | ".join(parts))
            return redirect(request.path)

        created = updated = skipped = 0
        with transaction.atomic():
            if target in ("thisinh", "voting"):
                for r in rows:
                    ma = (r.get("maNV") or "").strip()
                    if not ma:
                        skipped += 1
                        continue

                    hoTen = (r.get("hoTen") or "").strip()
                    chiNhanh = (r.get("chiNhanh") or "").strip() or None
                    vung = (r.get("vung") or "").strip() or None
                    donVi = (r.get("donVi") or "").strip() or None
                    email = (r.get("email") or "").strip() or None
                    nhom = (r.get("nhom") or "").strip() or None
                    image_url = (r.get("image_url") or "").strip() or None

                    ts, is_created = ThiSinh.objects.update_or_create(
                        pk=ma,
                        defaults=dict(
                            hoTen=hoTen,
                            chiNhanh=chiNhanh,
                            vung=vung,
                            donVi=donVi,
                            email=email,
                            nhom=nhom,
                            image_url=image_url,
                        ),
                    )

                    if cuocthi_obj:
                        # liên kết tham gia cuộc thi
                        try:
                            ThiSinhCuocThi.objects.get_or_create(thiSinh=ts, cuocThi=cuocthi_obj)
                        except Exception as e:
                            messages.warning(request, f"Lỗi khi tạo quan hệ Cuộc thi cho {ma}: {e}")

                        # nếu import 'voting' -> thêm vào ThiSinhVoting
                        if target == "voting":
                            try:
                                from core.models import ThiSinhVoting
                                ThiSinhVoting.objects.get_or_create(thiSinh=ts, cuocThi=cuocthi_obj)
                            except Exception as e:
                                messages.warning(request, f"Lỗi thêm Voting cho {ma}: {e}")

                    created += int(is_created)
                    updated += int(not is_created)

            else:  # giamkhao
                for r in rows:
                    ma = (r.get("maNV") or "").strip()
                    if not ma:
                        skipped += 1
                        continue
                    gk, is_created = GiamKhao.objects.update_or_create(
                        pk=ma,
                        defaults=dict(
                            hoTen=(r.get("hoTen") or "").strip(),
                            email=(r.get("email") or "").strip(),
                            role="JUDGE",
                        ),
                    )
                    created += int(is_created)
                    updated += int(not is_created)

        messages.success(request, f"Import xong: thêm {created}, cập nhật {updated}, bỏ qua {skipped}.")
        return redirect(request.path)

    # GET
    return render(
        request,
        "importer/index.html",
        {
            "cuocthi_list": CuocThi.objects.all().values("ma", "tenCuocThi").order_by("ma"),
            "preselected_ma": preselected_ma,
        }
        
)
def _extract_manv_from_filename_stem(stem: str) -> str | None:
    """
    Nhận vào phần tên file không có đuôi (stem) và trả về maNV nếu tìm được.
    Hỗ trợ tên file dạng:
      - 00041009 - Đỗ Văn Chiến - PSG02USR
      - 00053160_Nguyễn Trung Kiên_THN05USR2
      - avatar_00041009_do_van_chien
    Ưu tiên bắt chuỗi 8 chữ số (giữ được số 0 đầu).
    """
    if not stem:
        return None

    s = stem.strip()

    # 1) Ưu tiên: đúng 8 chữ số nằm độc lập trong chuỗi
    m = re.search(r"(?<!\d)(\d{8})(?!\d)", s)
    if m:
        return m.group(1)

    # 2) Fallback: lấy cụm số đầu chuỗi (>=6 số) nếu bạn có maNV không đúng 8 số
    m2 = re.match(r"^\s*(\d{6,})", s)
    if m2:
        return m2.group(1)

    return None


@judge_required
def upload_avatars_view(request):
    """
    - Upload nhiều ảnh avatar.
    - Tên file = maNV (không phân biệt hoa/thường), ví dụ: NV001.jpg
    - Lưu: MEDIA_ROOT/thisinh/<maNV>.<ext>
    - Cập nhật ThiSinh.image_url ứng với URL của file đã lưu.
    - Tuỳ chọn: chỉ cập nhật cho danh sách Voting của 1 cuộc thi (checkbox 'only_voting' + chọn 'maCT').
    """
    if request.method == "POST":
        files = request.FILES.getlist("images")
        selected_ma_ct = (request.POST.get("maCT") or "").strip()
        only_voting = bool(request.POST.get("only_voting"))

        if not files:
            messages.error(request, "Vui lòng chọn ít nhất 1 ảnh.")
            return redirect(request.path)

        # Tập mã thí sinh hợp lệ nếu chỉ cho danh sách Voting
        voting_ma_set_lower = None
        if only_voting:
            from core.models import ThiSinhVoting
            qs = ThiSinhVoting.objects.all()
            if selected_ma_ct:
                qs = qs.filter(cuocThi__ma=selected_ma_ct)
            voting_ma_set_lower = set(qs.values_list("thiSinh__maNV", flat=True))
            voting_ma_set_lower = {m.lower() for m in voting_ma_set_lower}

        updated = 0
        not_found = []
        skipped = 0
        skipped_not_in_voting = []

        for f in files:
            original_name = f.name or ""
            base_name = os.path.basename(original_name)
            stem, ext = os.path.splitext(base_name)

            ext = ext.lower()
            ma = _extract_manv_from_filename_stem(stem)

            if ext not in [".jpg", ".jpeg", ".png"]:
                skipped += 1
                continue
            if not ma:
                skipped += 1
                continue

            # Nếu bắt buộc thuộc danh sách voting
            if voting_ma_set_lower is not None and ma.lower() not in voting_ma_set_lower:
                skipped_not_in_voting.append(ma)
                continue

            # Tìm thí sinh
            try:
                ts = ThiSinh.objects.get(maNV__iexact=ma)
            except ThiSinh.DoesNotExist:
                not_found.append(ma)
                continue

            filename = f"{ts.maNV}{ext}"
            upload_path = os.path.join("thisinh", filename)

            if default_storage.exists(upload_path):
                default_storage.delete(upload_path)

            saved_path = default_storage.save(upload_path, f)

            try:
                url = default_storage.url(saved_path)
            except Exception:
                url = settings.MEDIA_URL + saved_path

            ts.image_url = url
            ts.save(update_fields=["image_url"])
            updated += 1

        if updated:
            messages.success(request, f"Đã cập nhật ảnh cho {updated} thí sinh.")
        if skipped:
            messages.warning(request, f"Bỏ qua {skipped} tệp không hợp lệ (sai định dạng hoặc thiếu mã).")
        if skipped_not_in_voting:
            messages.warning(
                request,
                "Bỏ qua (không thuộc danh sách Voting): " + ", ".join(sorted(set(skipped_not_in_voting))),
            )
        if not_found:
            messages.warning(
                request,
                "Không tìm thấy thí sinh cho các mã: " + ", ".join(sorted(set(not_found))),
            )

        return redirect(request.path)

    # GET: luôn trả về template
    return render(
        request,
        "importer/upload_avatars.html",
        {
            "cuocthi_list": CuocThi.objects.all().values("ma", "tenCuocThi").order_by("ma"),
        },
    )

# ============================================================
# TỔ CHỨC / IMPORT TEMPLATE CHẤM CHO 1 BÀI THI
# ============================================================

@judge_required
def organize_view(request):
    """
    Action 'config_template_upload':
      - Excel gồm cột Section, Item, Điểm (Max), Note (tuỳ chọn).
      - Tạo lại toàn bộ cấu trúc Section / Item cho BaiThi chỉ định.
    """
    if request.method == "POST":
        action = (request.POST.get("action") or "").strip()

        if action == "config_template_upload":
            btid = request.POST.get("baiThi_id")
            f = request.FILES.get("template_file")
            if not btid or not f:
                messages.error(request, "Thiếu Bài thi hoặc tệp Excel.")
                return redirect(request.path)

            try:
                bai_thi = BaiThi.objects.get(pk=btid)
            except BaiThi.DoesNotExist:
                messages.error(request, "Không tìm thấy Bài thi.")
                return redirect(request.path)

            try:
                wb = load_workbook(f, read_only=True, data_only=True)
                ws = wb.active
                rows = list(ws.iter_rows(values_only=True))
            except Exception as e:
                messages.error(request, f"Lỗi đọc Excel: {e}")
                return redirect(request.path)

            if not rows:
                messages.error(request, "Tệp rỗng.")
                return redirect(request.path)

            header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]

            def idx(*names):
                for name in names:
                    if name in header:
                        return header.index(name)
                return None

            idx_section = idx("section", "mục lớn", "muc lon", "phan", "phần")
            idx_item    = idx("item", "mục nhỏ", "muc nho", "bài", "bai")
            idx_max     = idx("điểm", "diem", "max", "điểm tối đa", "diem toi da")
            idx_note    = idx("note", "ghi chú", "ghi chu")

            if idx_section is None or idx_max is None:
                messages.error(request, "Thiếu cột bắt buộc: Section và Điểm tối đa.")
                return redirect(request.path)

            BaiThiTemplateSection.objects.filter(baiThi=bai_thi).delete()

            section_map = {}   # title -> (section_obj, next_item_stt)
            next_section_stt = 1
            created_items = 0

            for r in rows[1:]:
                if r is None:
                    continue

                sec_title = (r[idx_section] if idx_section is not None and idx_section < len(r) else None)
                item_text = (r[idx_item] if idx_item is not None and idx_item < len(r) else None)
                max_val   = (r[idx_max] if idx_max is not None and idx_max < len(r) else None)
                note_text = (r[idx_note] if idx_note is not None and idx_note < len(r) else None)

                if not (sec_title or item_text or max_val or note_text):
                    continue

                sec_title = str(sec_title).strip() if sec_title is not None else ""
                item_text = str(item_text).strip() if item_text is not None else ""
                note_text = str(note_text).strip() if note_text is not None else ""

                try:
                    max_score = int(float(max_val)) if max_val is not None and str(max_val).strip() != "" else 0
                except Exception:
                    messages.error(request, f"Điểm tối đa không hợp lệ ở dòng có mục: '{item_text or sec_title}'.")
                    return redirect(request.path)

                if sec_title not in section_map:
                    s = BaiThiTemplateSection.objects.create(
                        baiThi=bai_thi, stt=next_section_stt, title=sec_title or "Mục"
                    )
                    section_map[sec_title] = [s, 1]
                    next_section_stt += 1
                else:
                    s, _ = section_map[sec_title]

                s, next_item_stt = section_map[sec_title]
                BaiThiTemplateItem.objects.create(
                    section=s,
                    stt=next_item_stt,
                    content=(item_text or s.title),
                    max_score=max_score,
                    note=note_text or None,
                )
                section_map[sec_title][1] = next_item_stt + 1
                created_items += 1

            messages.success(
                request,
                f"Đã cập nhật mẫu chấm cho {bai_thi.tenBaiThi}: {len(section_map)} mục lớn, {created_items} mục nhỏ."
            )
            return redirect(request.path)

        # Các action khác có thể bổ sung sau
        messages.error(request, "Hành động chưa được hỗ trợ.")
        return redirect(request.path)

    # GET: hiển thị trang tổ chức
    cuoc_this = (
        CuocThi.objects
        .prefetch_related(
            Prefetch("vong_thi", queryset=VongThi.objects.prefetch_related(
                Prefetch("bai_thi", queryset=BaiThi.objects.prefetch_related("time_rules", "template_sections__items"))
            ))
        )
        .order_by("ma")
    )
    return render(request, "organize/index.html", {"cuoc_this": cuoc_this})
