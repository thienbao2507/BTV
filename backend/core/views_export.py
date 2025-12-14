# core/views_export.py
from __future__ import annotations
from django.http import HttpResponse
from django.shortcuts import render, get_object_or_404
from django.db.models import Avg, Max
from decimal import Decimal, ROUND_HALF_UP
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side  # <- thêm Border, Side
from .models import CuocThi, VongThi, BaiThi, ThiSinh, PhieuChamDiem
from .models import SpecialRoundPairMember

# --- helpers cho thời gian ---
def _pick_time_value(obj):
    """
    Trích xuất thời gian (giây) từ một record PhieuChamDiem.
    Hỗ trợ linh hoạt nhiều tên field khác nhau.
    Trả về int(giây) hoặc None nếu không có.
    """
    CANDIDATES = ["thoiGian", "thoiGianGiay", "time_seconds", "time", "duration", "tongThoiGian"]
    for k in CANDIDATES:
        if hasattr(obj, k):
            v = getattr(obj, k)
            if v is None: 
                continue
            try:
                # chấp nhận float/decimal → ép int giây
                return int(round(float(v)))
            except Exception:
                pass
    return None

def _fmt_mmss(seconds: int | None) -> str:
    if seconds is None:
        return ""
    m, s = divmod(max(0, int(seconds)), 60)
    return f"{m:02d}:{s:02d}"

def _score_type(bt) -> str:
    v = getattr(bt, "phuongThucCham", None)
    if v is None:
        return "POINTS"
    s = str(v).strip().upper()
    if s in {"TIME", "2"}:
        return "TIME"
    if s in {"TEMPLATE", "1"}:
        return "TEMPLATE"
    return "POINTS"

def _build_columns(ct: CuocThi):
    vong_ids = VongThi.objects.filter(cuocThi=ct).values_list("id", flat=True)
    bai_qs = (
        BaiThi.objects
        .filter(vongThi_id__in=vong_ids)
        .select_related("vongThi")
        .prefetch_related("time_rules", "template_sections__items")
        .order_by("vongThi_id", "id")
    )

    cols = []
    for b in bai_qs:
        # tính max điểm để hiện trên header (giữ nguyên logic cũ)
        if _score_type(b) == "TIME":
            rules = list(b.time_rules.all()) if hasattr(b, "time_rules") else []
            b_max = max([r.score for r in rules], default=0)
        elif _score_type(b) == "TEMPLATE":
            b_max = sum(i.max_score for s in b.template_sections.all() for i in s.items.all())
        else:
            b_max = b.cachChamDiem

        # 1) Cột điểm (giữ tiêu đề 2 dòng để JS nhận diện là cột điểm)
        cols.append({
            "id": b.id,
            "code": b.ma,
            "kind": "score",
            "title": f"{b.vongThi.tenVongThi}\n{b.tenBaiThi}",
            "max": b_max,
        })
        # 2) Cột thời gian đi kèm (đặt ngay sau cột điểm)
        cols.append({
            "id": b.id,
            "code": b.ma,
            "kind": "time",
            # tiêu đề rõ ràng: “Thời gian (BTxxx)” – có thể rút gọn nếu muốn
            "title": f"Thời gian",
            "max": None,
        })

    titles = [c["title"] for c in cols]   # chỉ tiêu đề phần bài thi
    return cols, titles


def _flatten(ct: CuocThi):
    cols_meta, titles_per_exam = _build_columns(ct)

    # Info columns bạn đang dùng
    info_titles = ['Đơn vị', 'Chi nhánh', 'Vùng', 'Nhóm', 'Email']

    # Header đầy đủ
    columns = ['STT', 'Mã NV', 'Họ tên'] + info_titles + titles_per_exam + ['Tổng', 'Tổng thời gian']

    # ==== Map điểm trung bình theo (maNV, baiThi_id)
    score_qs = (
        PhieuChamDiem.objects
        .filter(cuocThi=ct)
        .values("thiSinh__maNV", "baiThi_id")
        .annotate(avg=Avg("diem"))
    )
    score_map = {(r["thiSinh__maNV"], r["baiThi_id"]): (float(r["avg"]) if r["avg"] is not None else "") for r in score_qs}

    # ==== Map thời gian ưu tiên MIN theo (maNV, baiThi_id)
    all_phieu = list(PhieuChamDiem.objects.filter(cuocThi=ct).select_related("thiSinh", "baiThi"))
    time_map = {}
    for p in all_phieu:
        key = (getattr(p.thiSinh, "maNV", None), getattr(p.baiThi, "id", None))
        if key[0] is None or key[1] is None:
            continue
        t = _pick_time_value(p)  # giây hoặc None
        if t is None:
            continue
        cur = time_map.get(key)
        if (cur is None) or (t < cur):
            time_map[key] = t

    ts_qs = ThiSinh.objects.filter(cuocThi=ct).order_by("maNV").distinct()
    def _sv(x): return "" if x is None else str(x)

    # ==== TẬP vòng/bài/thi-sinh thuộc vòng đặc biệt ====
    # 1) Ưu tiên cờ is_special_bonus_round; nếu không có, fallback theo các cặp tồn tại
    vt_special_ids = list(
        VongThi.objects.filter(cuocThi=ct, is_special_bonus_round=True).values_list("id", flat=True)
    )
    if not vt_special_ids:
        vt_special_ids = list(
            SpecialRoundPairMember.objects
            .filter(pair__vongThi__cuocThi=ct)
            .values_list("pair__vongThi_id", flat=True)
            .distinct()
        )

    bt_special_ids = set(
        BaiThi.objects.filter(vongThi_id__in=vt_special_ids).values_list("id", flat=True)
    )

    special_members_ma = set(
        SpecialRoundPairMember.objects
        .filter(pair__vongThi_id__in=vt_special_ids)
        .values_list("thiSinh__maNV", flat=True)
        .distinct()
    )

    special_sort_active = False
    if vt_special_ids and bt_special_ids and special_members_ma:
        special_sort_active = PhieuChamDiem.objects.filter(
            cuocThi=ct,
            vongThi_id__in=vt_special_ids,
            baiThi_id__in=bt_special_ids,
        ).exists()

    # ==== Gom dữ liệu từng thí sinh ====
    data = []
    bt_ids_in_order = [c["id"] for c in cols_meta if c.get("kind") == "score"]

    for ts in ts_qs:
        row = [
            None,  # STT, sẽ gán sau
            _sv(getattr(ts, "maNV", "")),
            _sv(getattr(ts, "hoTen", "")),
            _sv(getattr(ts, "donVi", "")),
            _sv(getattr(ts, "chiNhanh", "")),
            _sv(getattr(ts, "vung", "")),
            _sv(getattr(ts, "nhom", "")),
            _sv(getattr(ts, "email", "")),
        ]

        total_score = 0.0
        total_time_sec = 0
        has_any_time = False

        # Vừa build row, vừa tính tổng
        for bt_id in bt_ids_in_order:
            sc = score_map.get((ts.maNV, bt_id), "")
            row.append(sc)
            if isinstance(sc, (int, float, Decimal)):
                total_score += float(sc)

            tm_seconds = time_map.get((ts.maNV, bt_id))
            row.append(_fmt_mmss(tm_seconds))
            if tm_seconds is not None:
                has_any_time = True
                total_time_sec += tm_seconds

        # Cột Tổng
        row.append(
            int(Decimal(total_score).quantize(0, rounding=ROUND_HALF_UP))
        )

        # Cột Tổng thời gian
        row.append(_fmt_mmss(total_time_sec) if has_any_time else "")

        # ==== Tính tổng điểm ở các bài "vòng đặc biệt" để suy ra group ====
        # ==== Tính nhóm vòng đặc biệt: Winner(2) > Loser(1) > NoScore(0) ====
        sp_total = 0.0
        sp_has_score = False

        if ts.maNV in special_members_ma and bt_special_ids:
            for bt_id in bt_special_ids:
                sc_sp = score_map.get((ts.maNV, bt_id), "")
                if isinstance(sc_sp, (int, float, Decimal)):
                    sp_has_score = True
                    sp_total += float(sc_sp)

        if special_sort_active and (ts.maNV in special_members_ma):
            if sp_has_score:
                special_group = 2 if sp_total > 0 else 1
            else:
                special_group = 0
        else:
            special_group = 0



        data.append({
            "ts": ts,
            "row": row,
            "total_score": float(total_score),
            "total_time_sec": (total_time_sec if has_any_time else None),
            "__special_group": special_group,
        })

    # ==== Sort theo NHÓM → TỔNG → THỜI GIAN → MÃ NV ====
    def _time_key(seconds):
        return seconds if seconds is not None else float("inf")

    data.sort(
        key=lambda d: (
            -int(d.get("__special_group", 0)),
            -float(d["total_score"]),
            _time_key(d["total_time_sec"]),
            _sv(getattr(d["ts"], "maNV", "")),
        )
    )


    # Gán STT + build special_groups cùng thứ tự rows
    rows = []
    special_groups = []
    for idx, item in enumerate(data, start=1):
        item["row"][0] = idx
        rows.append(item["row"])
        special_groups.append(int(item.get("__special_group", 0)))

    # Trả thêm special_groups để front-end có thể ưu tiên nhóm khi sort cột "Tổng"
    return columns, rows, special_groups




# (Giữ export_page như cũ)

# (Tuỳ ý: có thể xoá export_csv và route của nó)
# def export_csv(...):  # <-- BỎ KHI KHÔNG DÙNG NỮA
#     ...

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from io import BytesIO
def export_xlsx(request):
    ct_id = request.GET.get("ct")
    ct = get_object_or_404(CuocThi, id=ct_id)

    use_visible = (request.method == "POST")
    if use_visible:
        # Nhận payload từ frontend
        import json
        try:
            payload = json.loads(request.body.decode("utf-8"))
            columns = payload.get("columns") or []
            rows = payload.get("rows") or []
            kinds = payload.get("col_kinds") or ["info"] * len(columns)
        except Exception:
            # fallback sang full nếu payload lỗi
            columns, rows, _special_groups = _flatten(ct)
            kinds = ["info"] * len(columns)
    else:
        columns, rows, _special_groups = _flatten(ct)

        info_count = 3 + 5
        kinds = ["info"] * len(columns)

        total_cols = len(columns)
        # ít nhất: info (8 cột) + 2 cột tổng
        if total_cols >= info_count + 2:
            j = info_count
            last_score_idx = total_cols - 2  # index của cột "Tổng"

            # Các cặp Điểm/Thời gian theo từng bài thi
            while j < last_score_idx:
                kinds[j] = "score"   # cột Điểm
                j += 1
                if j < last_score_idx:
                    kinds[j] = "time"   # cột Thời gian
                    j += 1

            # Hai cột cuối: "Tổng" + "Tổng thời gian"
            kinds[-2] = "score"  # Tổng
            kinds[-1] = "time"   # Tổng thời gian
        else:
            # fallback an toàn nếu cấu trúc columns khác kỳ vọng
            j = info_count
            while j < total_cols:
                kinds[j] = "score"
                j += 1



    wb = Workbook()
    ws = wb.active
    ws.title = f"{ct.ma}"

    # ==== Header: KHÔNG in đậm ====
    ws.append(columns)
    for c in range(1, len(columns)+1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True, size=12)  # KHÔNG in đậm
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ==== Body: dùng rows đã có (full hoặc visible) ====
    for r in rows:
        ws.append(r)

    # ==== Tô màu theo CỘT ====
    fill_info  = PatternFill(fill_type="solid", start_color="FFEAF4FF", end_color="FFEAF4FF")  # xanh nhạt
    fill_score = PatternFill(fill_type="solid", start_color="FFFFF5E6", end_color="FFFFF5E6")  # vàng nhạt

    max_row = ws.max_row
    max_col = ws.max_column
    for j in range(1, max_col+1):
        kind = kinds[j-1] if (j-1) < len(kinds) else "info"
        fill = fill_score if kind == "score" else fill_info
        for i in range(1, max_row+1):
            ws.cell(row=i, column=j).fill = fill


    # ... sau khi append header + body và tô màu, tính sẵn:
    max_row = ws.max_row
    max_col = ws.max_column

    # ==== Border mảnh cho toàn bộ ô ====
    thin = Side(style="thin", color="FF000000")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

    for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col), start=1):
        for cell in row:
            cell.border = border_all
            # Font: header (r1) 12pt, body 11pt, không đậm (đúng yêu cầu “không in đậm”)
            if r_idx == 1:
                cell.font = Font(name="Times new roman", size=12, bold=True)
            else:
                cell.font = Font(name="Times new roman", size=11, bold=False)
            # Alignment
            cell.alignment = Alignment(
                vertical="center",
                wrap_text=True,
                horizontal=cell.alignment.horizontal if cell.alignment else "left"
            )

    # (giữ nguyên) Freeze 3 cột + 1 hàng tiêu đề
    ws.freeze_panes = "E2"

    # ==== Auto width (rộng hơn một chút) ====
    for i, col in enumerate(columns, start=1):
        maxlen = len(str(col)) if col is not None else 0
        for r in rows:
            v = r[i-1] if i-1 < len(r) else ""
            l = len(str(v)) if v is not None else 0
            if l > maxlen:
                maxlen = l
        # padding rộng hơn: +4, tối thiểu 12, tối đa 60
        ws.column_dimensions[get_column_letter(i)].width = max(12, min(maxlen + 4, 60))


    # Giữ freeze panes cũ (3 cột trái + 1 hàng tiêu đề)
    ws.freeze_panes = "E2"

    bio = BytesIO(); wb.save(bio); bio.seek(0)
    resp = HttpResponse(
        bio.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    # Tên file: nếu POST visible thì đổi chút cho phân biệt
    fname = f'export_{ct.ma}.xlsx' if not use_visible else f'export_{ct.ma}.xlsx'
    resp["Content-Disposition"] = f'attachment; filename="{fname}"'
    return resp
def export_page(request):
    ct_id = request.GET.get("ct")
    ct = get_object_or_404(CuocThi, id=ct_id)

    # Chỉ lấy các cuộc thi đang bật
    active_cts = CuocThi.objects.filter(trangThai=True).order_by("ma", "tenCuocThi")

    columns, rows, special_groups = _flatten(ct)
    return render(request, "export/index.html", {
        "contest": ct,
        "columns": columns,
        "rows": rows,
        "special_groups": special_groups,   # NEW
        "active_cts": active_cts,
    })
# --- FINAL EXPORT (Chung Kết) ---
from django.db.models import Avg, Sum
from django.http import JsonResponse
from .models import CuocThi, VongThi, BaiThi, ThiSinh, PhieuChamDiem, BattleVote, BGDScore


def _find_chung_ket():
    """
    Tìm cuộc thi 'Chung Kết' theo nhiều biến thể: 'Chung Kết' / 'Chung Ket' (không dấu).
    Ưu tiên __iexact, fallback bản không dấu thô.
    """
    ct = CuocThi.objects.filter(tenCuocThi__iexact="Chung Kết").first()
    if not ct:
        ct = CuocThi.objects.filter(tenCuocThi__iexact="Chung Ket").first()
    return ct

def _final_columns_and_rows(ct: CuocThi):
    """
    Trả về (columns, rows) cho trang Export Chung Kết:
    - Cột info như export thường: STT, Mã NV, Họ tên, Đơn vị, Chi nhánh, Vùng, Nhóm, Email
    - Cột điểm: Tổng điểm (vòng Chung kết), Đối kháng (sao TB, 1 số thập phân)
    """
    info_titles = ['STT', 'Mã NV', 'Họ tên', 'Đơn vị', 'Chi nhánh', 'Vùng', 'Nhóm', 'Email']
    columns = info_titles + ['Đối kháng', 'Tim','Soán ngôi','Tổng điểm']



    # 1) Xác định Vòng “Chung Kết” (nếu không tìm được thì lấy tất cả vòng của CT này)
    vt_ck = VongThi.objects.filter(cuocThi=ct, tenVongThi__iexact="Chung Kết")
    if not vt_ck.exists():
        vt_ck = VongThi.objects.filter(cuocThi=ct)
    vt_ids = list(vt_ck.values_list("id", flat=True))

    # 2) Các bài thi thuộc vòng CK
    bt_ids = list(BaiThi.objects.filter(vongThi_id__in=vt_ids).values_list("id", flat=True))

    # 3) Map tổng điểm vòng CK cho từng thí sinh
    #    (gộp trung bình theo bài, rồi SUM các bài)
    score_qs = (
        PhieuChamDiem.objects
        .filter(cuocThi=ct, baiThi_id__in=bt_ids)
        .values("thiSinh__maNV", "baiThi_id")
        .annotate(avg=Avg("diem"))
    )
    # tích lũy SUM(avg) theo thí sinh
    total_by_ma = {}
    for r in score_qs:
        ma = r["thiSinh__maNV"]
        total_by_ma[ma] = total_by_ma.get(ma, 0.0) + float(r["avg"] or 0.0)

    # 4) Đối kháng: trung bình sao theo BattleVote cho CT này
    #    BattleVote.entry -> ThiSinhCapThiDau -> pair -> cuocThi
    #    Lấy TB sao theo thiSinh (entry.thiSinh.maNV)
    battle_qs = (
        BattleVote.objects
        .filter(entry__pair__cuocThi=ct)
        .values("entry__thiSinh__maNV")
        .annotate(avg=Avg("stars"))
    )
    stars_by_ma = {r["entry__thiSinh__maNV"]: (float(r["avg"]) if r["avg"] is not None else None)
                   for r in battle_qs}

    # NEW: đếm số "Tim" (♥) theo thí sinh trong CK
    from django.db.models import Case, When, IntegerField, Sum
    heart_qs = (
        BattleVote.objects
        .filter(entry__pair__cuocThi=ct)
        .values("entry__thiSinh__maNV")
        .annotate(hearts=Sum(Case(When(heart=True, then=1), default=0, output_field=IntegerField())))
    )
    hearts_by_ma = {r["entry__thiSinh__maNV"]: int(r["hearts"] or 0) for r in heart_qs}

    # 4b) Soán ngôi: điểm trung bình BGDScore theo thí sinh cho cuộc thi này
    #     Mỗi BGD chỉ còn 1 dòng (điểm mới nhất) cho thí sinh trong bảng BGDScore
    avg_qs = (
        BGDScore.objects
        .filter(cuocThi=ct)
        .values("thiSinh__maNV")
        .annotate(avg=Avg("diem"))
    )

    soan_by_ma = {
        r["thiSinh__maNV"]: (float(r["avg"]) if r["avg"] is not None else None)
        for r in avg_qs
    }



    # 5) Duyệt thí sinh của CT & build rows
    ts_qs = ThiSinh.objects.filter(cuocThi=ct).order_by("maNV").distinct()

    def _sv(x): return "" if x is None else str(x)

    # Gom dữ liệu thô để sắp xếp trước
    data = []
    for ts in ts_qs:
        # Đối kháng
        sao = stars_by_ma.get(ts.maNV, None)
        sao_fmt = (str(int(Decimal(str(sao)).quantize(0, rounding=ROUND_HALF_UP))) if sao is not None else "")
        sao_val = int(Decimal(str(sao or 0)).quantize(0, rounding=ROUND_HALF_UP))

        # Tim
        tim = hearts_by_ma.get(ts.maNV, 0)

        # Soán ngôi (AVG điểm BGD)
        soan = soan_by_ma.get(ts.maNV, None)
        soan_fmt = (str(int(Decimal(str(soan)).quantize(0, rounding=ROUND_HALF_UP))) if soan is not None else "")
        soan_val = int(Decimal(str(soan or 0)).quantize(0, rounding=ROUND_HALF_UP))

        # Tổng điểm hiển thị = Soán ngôi + Đối kháng
        total_display = int(Decimal(soan_val + sao_val).quantize(0, rounding=ROUND_HALF_UP))

        data.append({
            "ts": ts,
            "sao_fmt": sao_fmt,
            "sao_val": sao_val,
            "tim": tim,
            "soan_fmt": soan_fmt,
            "soan_val": soan_val,
            "total": total_display,
        })

    # Sort mặc định:
    # 1) Tổng điểm giảm dần
    # 2) Nếu bằng nhau -> Tim giảm dần
    # 3) Cuối cùng sort theo Mã NV cho ổn định
    data.sort(
        key=lambda d: (
            -int(d["total"]),
            -int(d["tim"]),
            _sv(getattr(d["ts"], "maNV", "")),
        )
    )

    # Build rows sau khi đã sort
    rows = []
    for idx, item in enumerate(data, start=1):
        ts = item["ts"]
        row = [
            idx,
            _sv(getattr(ts, "maNV", "")),
            _sv(getattr(ts, "hoTen", "")),
            _sv(getattr(ts, "donVi", "")),
            _sv(getattr(ts, "chiNhanh", "")),
            _sv(getattr(ts, "vung", "")),
            _sv(getattr(ts, "nhom", "")),
            _sv(getattr(ts, "email", "")),
            item["sao_fmt"],     # Đối kháng
            item["tim"],         # Tim
            item["soan_fmt"],    # Soán ngôi
            item["total"],       # Tổng điểm = Soán ngôi + Đối kháng
        ]
        rows.append(row)

    return columns, rows

def export_final_page(request):
    """
    Trang web Export Chung Kết (bảng Excel-like).
    Yêu cầu truyền ct=<id> giống export thường để tránh nhầm cuộc thi.
    """
    ct_id = request.GET.get("ct")
    if not ct_id:
        return render(request, "export/index.html", {
            "contest": None,
            "columns": [],
            "rows": [],
            "FROZEN_COUNT": 3,
            "final_mode": True,
            "error": "Thiếu tham số ?ct=<id> cho Export Chung Kết.",
        })

    ct = get_object_or_404(CuocThi, id=ct_id)

    columns, rows = _final_columns_and_rows(ct)
    return render(request, "export/index.html", {
        "contest": ct,
        "columns": columns,
        "rows": rows,
        "FROZEN_COUNT": 3,
        "final_mode": True,
    })


from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from io import BytesIO

def export_final_xlsx(request):
    """
    Xuất XLSX cho Chung Kết (giống export-xlsx nhưng chỉ 2 cột điểm).
    Bắt buộc truyền ?ct=<id> để chọn đúng cuộc thi.
    """
    ct_id = request.GET.get("ct")
    if not ct_id:
        return HttpResponse("Thiếu tham số ?ct=<id> cho Export Chung Kết.", status=400)

    ct = get_object_or_404(CuocThi, id=ct_id)

    columns, rows = _final_columns_and_rows(ct)

    wb = Workbook()
    ws = wb.active
    ws.title = f"{ct.ma}"

    # Header
    ws.append(columns)
    for c in range(1, len(columns)+1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Body
    for r in rows:
        ws.append(r)

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    resp = HttpResponse(out.read(),
                        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = f'attachment; filename="export_chungket_{ct.ma}.xlsx"'
    return resp
